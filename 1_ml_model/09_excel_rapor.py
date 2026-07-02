"""
1_ML_Model/09_excel_rapor.py — Çok Sekmeli Excel Raporu
=========================================================
Çıktı: 1_ML_Model/rapor_ml_model.xlsx
Sekmeler:
  Genel Özet, Boş Geçmiş, Kaçırılan, Boşa Giden Tahminler,
  Doğru Tahminler, Veri Kalitesi Şüpheli, İş Anomalileri
Her sekmede Excel outline (+/-) ile cari/ürün gruplama

Çalıştır: python 1_ML_Model/09_excel_rapor.py
"""

import pathlib, sys, warnings
warnings.filterwarnings("ignore")

KOK = pathlib.Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(KOK))
from veri_yukleme import yukle_tahmin

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CALISMA_DIR = pathlib.Path(__file__).parent.resolve()
ARA_DIR     = CALISMA_DIR / "ara_ciktilar"
CIKTI_EXCEL = CALISMA_DIR / "rapor_ml_model.xlsx"

MAX_SATIR_EXCEL = 50_000   # Excel sekmesine yazılacak maksimum satır

# Excel stil sabitleri
BASLIK_DOLGU  = PatternFill("solid", fgColor="1A535C")
BASLIK_FONT   = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_DOLGU   = PatternFill("solid", fgColor="EDF7F6")
UYARI_DOLGU   = PatternFill("solid", fgColor="FFE0B2")  # büyük hata için turuncu
GRUPLAMA_DOLGU= PatternFill("solid", fgColor="B2DFDB")  # grup başlık satırı
INCE_CERCEVE  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
GRUP_FONT     = Font(bold=True, size=10, color="1A535C")

print("=" * 60)
print("EXCEL RAPORU OLUŞTURULUYOR")
print("=" * 60)

# ── Verileri Yükle ─────────────────────────────────────────────────────────
tahmin_temiz = yukle_tahmin()
panel        = pd.read_parquet(ARA_DIR / "panel_with_predictions.parquet")
panel["hata"] = panel["yeni_tahmin_miktar"] - panel["ToplamMiktar"]
metrikler    = pd.read_parquet(ARA_DIR / "metrik_karsilastirma.parquet")

try:
    vk_supheli = pd.read_parquet(ARA_DIR / "veri_kalitesi_supheli.parquet")
except: vk_supheli = pd.DataFrame()
try:
    mukerrer = pd.read_parquet(ARA_DIR / "mukerrer_kayitlar.parquet")
except: mukerrer = pd.DataFrame()
try:
    anomali_yuzdelik = pd.read_parquet(ARA_DIR / "anomali_is_yuzdelik.parquet")
except: anomali_yuzdelik = pd.DataFrame()
try:
    churn = pd.read_parquet(ARA_DIR / "anomali_churn.parquet")
except: churn = pd.DataFrame()
try:
    sistematik_cari = pd.read_parquet(ARA_DIR / "anomali_sistematik_cari.parquet")
except: sistematik_cari = pd.DataFrame()

# ── Confusion Matrix Sınıfları ─────────────────────────────────────────────
y_gercek  = panel["siparis_var"].values
y_yeni    = panel["yeni_siparis_var"].values
y_son_fin = panel.get("final_siparis", pd.Series(0, index=panel.index)).values

# TP, FN, FP — yeni_tahmin üzerinden
tp_mask = (y_gercek == 1) & (y_yeni == 1)
fn_mask = (y_gercek == 1) & (y_yeni == 0)
fp_mask = (y_gercek == 0) & (y_yeni == 1)

df_tp  = panel[tp_mask].copy()
df_fn  = panel[fn_mask].copy()
df_fp  = panel[fp_mask].copy()

# Hız ve dosya boyutu için sınırlandırma
print(f"  TP (Doğru Tahmin)      : {len(df_tp):,} (en büyük hatalı {MAX_SATIR_EXCEL:,} satır yazılacak)")
if not df_tp.empty:
    df_tp["abs_hata"] = df_tp["hata"].abs()
    df_tp = df_tp.nlargest(MAX_SATIR_EXCEL, "abs_hata").drop(columns=["abs_hata"])

print(f"  FN (Kaçırılan)         : {len(df_fn):,}")
print(f"  FP (Boşa Giden)        : {len(df_fp):,}")

# Cold-start: geçmiş hiç yok
adi_cv2 = pd.read_parquet(ARA_DIR / "adi_cv2.parquet")
tahmin_temiz["key"] = tahmin_temiz["cari_kod"].astype(str) + "_" + tahmin_temiz["stok_kod"].astype(str)
adi_cv2_keys = set(adi_cv2["CariKod"].astype(str) + "_" + adi_cv2["StockKod"].astype(str))
gormemis = tahmin_temiz[~tahmin_temiz["key"].isin(adi_cv2_keys)].copy().drop(columns=["key"])
tahmin_temiz = tahmin_temiz.drop(columns=["key"], errors="ignore")
print(f"  Boş Geçmiş (cold-start): {len(gormemis):,} (en yüksek tahminli {MAX_SATIR_EXCEL:,} satır yazılacak)")
if not gormemis.empty:
    gormemis = gormemis.nlargest(MAX_SATIR_EXCEL, "final_tahmin")

# İş anomalileri de limitlenmeli
try:
    anomali_yuzdelik = pd.read_parquet(ARA_DIR / "anomali_is_yuzdelik.parquet")
    if not anomali_yuzdelik.empty:
        anomali_yuzdelik["abs_hata"] = anomali_yuzdelik["hata"].abs()
        anomali_yuzdelik = anomali_yuzdelik.sort_values("abs_hata", ascending=False).head(20000).drop(columns=["abs_hata"])
except:
    anomali_yuzdelik = pd.DataFrame()


# ── Yardımcı Fonksiyonlar ──────────────────────────────────────────────────
def otomatik_genislik(ws, maks=55):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), maks))
        ws.column_dimensions[col_letter].width = min(max_len + 4, maks)

def bicimlendir_baslik(ws, satir=1):
    for cell in ws[satir]:
        cell.font      = BASLIK_FONT
        cell.fill      = BASLIK_DOLGU
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = INCE_CERCEVE
    ws.row_dimensions[satir].height = 28

def val_temizle(val):
    if pd.isnull(val):
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%d.%m.%Y")
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return round(float(val), 4)
    return val

def df_to_sheet_grouped(wb, sheet_name, df_kural, gruplama_kol="CariKod",
                        miktar_hata_kol=None, hata_esik=None):
    """
    DataFrame'i gruplama_kol'a göre Excel outline (+/-) ile yazar.
    Grup başlık satırı (özet) → detay satırlar gizlenebilir.
    """
    ws = wb.create_sheet(title=sheet_name)
    if df_kural is None or df_kural.empty:
        ws.append(["Bu sekme için kayıt bulunamadı."])
        ws["A1"].font = Font(italic=True, color="888888")
        return ws

    kolonlar = list(df_kural.columns)
    ws.append(kolonlar)
    bicimlendir_baslik(ws)

    # Gruplama
    if gruplama_kol not in df_kural.columns:
        # Gruplama kolonu yoksa düz yaz
        for i, row in enumerate(df_kural.itertuples(index=False, name=None), start=2):
            ws.append([val_temizle(v) for v in row])
    else:
        df_s = df_kural.sort_values(gruplama_kol).reset_index(drop=True)
        gruplar = df_s.groupby(gruplama_kol, sort=False)
        satir_no = 2
        for grup_key, grup_df in gruplar:
            # Grup başlık satırı
            grup_ozet = [f"▶ {gruplama_kol}: {grup_key}  ({len(grup_df)} kayıt)"]
            grup_ozet += [""] * (len(kolonlar) - 1)
            ws.append(grup_ozet)
            for cell in ws[satir_no]:
                cell.font      = GRUP_FONT
                cell.fill      = GRUPLAMA_DOLGU
                cell.alignment = Alignment(horizontal="left")
            satir_no += 1
            detay_baslangic = satir_no

            # Detay satırlar
            for i, row in enumerate(grup_df.itertuples(index=False, name=None)):
                ws.append([val_temizle(v) for v in row])
                # Büyük miktar hatası → turuncu vurgula
                if miktar_hata_kol and hata_esik:
                    try:
                        kol_idx = kolonlar.index(miktar_hata_kol) + 1
                        hata_val = ws.cell(row=satir_no, column=kol_idx).value
                        if hata_val is not None and abs(float(hata_val)) > hata_esik:
                            for cell in ws[satir_no]:
                                cell.fill = UYARI_DOLGU
                    except Exception:
                        pass
                # Zebra
                if i % 2 == 0:
                    for cell in ws[satir_no]:
                        if cell.fill.fgColor.rgb == "00000000":
                            cell.fill = ZEBRA_DOLGU
                satir_no += 1

            # Outline gruplama
            for r in range(detay_baslangic, satir_no):
                ws.row_dimensions[r].outline_level = 1
                ws.row_dimensions[r].hidden = False

    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = True
    otomatik_genislik(ws)
    return ws

# ── Workbook Oluştur ────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws_ozet = wb.active
ws_ozet.title = "Genel Özet"

# ── Genel Özet Sekmesi ───────────────────────────────────────────────────────
ws_ozet.merge_cells("A1:F1")
ws_ozet["A1"] = "ML MODEL — GENEL ÖZET ve 7 METRİK KARŞILAŞTIRMASI"
ws_ozet["A1"].font      = Font(bold=True, size=14, color="1A535C")
ws_ozet["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws_ozet.row_dimensions[1].height = 38

# Temel sayılar
ws_ozet.append([])
ws_ozet.append(["Metrik", "Değer"])
for cell in ws_ozet[ws_ozet.max_row]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4ECDC4")
ws_ozet.append(["Doğru Tahminler (TP)", len(df_tp)])
ws_ozet.append(["Kaçırılan (FN)", len(df_fn)])
ws_ozet.append(["Boşa Giden Tahminler (FP)", len(df_fp)])
ws_ozet.append(["Boş Geçmiş (Cold-Start)", len(gormemis)])
ws_ozet.append([])

# 7 Metrik karşılaştırma tablosu
ws_ozet.append(["7 METRİK KARŞILAŞTIRMA TABLOSU"])
ws_ozet[f"A{ws_ozet.max_row}"].font = Font(bold=True, size=12, color="1A535C")
ws_ozet.append([])

# Tablo başlığı
tablo_bas = ["Metrik",
             "Final Tahmin — Tüm", "Yeni Tahmin — Tüm",
             "Final Tahmin — Aktif", "Yeni Tahmin — Aktif"]
ws_ozet.append(tablo_bas)
for cell in ws_ozet[ws_ozet.max_row]:
    cell.font = BASLIK_FONT
    cell.fill = BASLIK_DOLGU
    cell.alignment = Alignment(horizontal="center")

for _, row in metrikler.iterrows():
    ws_ozet.append([
        row["Metrik"],
        round(row["Final Tahmin — Tüm"], 4),
        round(row["Yeni Tahmin — Tüm"], 4),
        round(row["Final Tahmin — Aksiyon Alınabilir"], 4),
        round(row["Yeni Tahmin — Aksiyon Alınabilir"], 4),
    ])

otomatik_genislik(ws_ozet)
ws_ozet.freeze_panes = "A2"

# ── Diğer Sekmeler ─────────────────────────────────────────────────────────
sekme_tanimi = [
    ("Boş Geçmiş",             gormemis[["cari_kod","stok_kod","urun_adi",
                                          "birim","oneri_hafta","final_tahmin"]]
                                .drop_duplicates() if not gormemis.empty else pd.DataFrame(),
     "cari_kod", None, None),
    ("Kaçırılan",              df_fn[["CariKod","StockKod","Hafta","ToplamMiktar",
                                       "yeni_tahmin_miktar","hata"]]
                                .rename(columns={"CariKod":"cari_kod","StockKod":"stok_kod"}),
     "cari_kod", None, None),
    ("Boşa Giden Tahminler",   df_fp[["CariKod","StockKod","Hafta",
                                       "yeni_tahmin_miktar","siparis_proba"]]
                                .rename(columns={"CariKod":"cari_kod","StockKod":"stok_kod"}),
     "cari_kod", None, None),
    ("Doğru Tahminler",        df_tp[["CariKod","StockKod","Hafta","ToplamMiktar",
                                       "yeni_tahmin_miktar","hata",
                                       "final_tahmin"]]
                                .rename(columns={"CariKod":"cari_kod","StockKod":"stok_kod"})
                                if "final_tahmin" in df_tp.columns else
                                df_tp[["CariKod","StockKod","Hafta","ToplamMiktar",
                                        "yeni_tahmin_miktar","hata"]]
                                .rename(columns={"CariKod":"cari_kod","StockKod":"stok_kod"}),
     "cari_kod", "hata", 50),
    ("Veri Kalitesi Şüpheli",  vk_supheli, "CariKod", None, None),
    ("İş Anomalileri",
     pd.concat([
         anomali_yuzdelik.assign(anomali_tipi="Yüzdelik Dışı") if not anomali_yuzdelik.empty else pd.DataFrame(),
         churn.assign(anomali_tipi="Churn Sinyali") if not churn.empty else pd.DataFrame(),
         sistematik_cari.assign(anomali_tipi="Sistematik Yön Hatası") if not sistematik_cari.empty else pd.DataFrame(),
     ], ignore_index=True),
     "CariKod" if "CariKod" in anomali_yuzdelik.columns else "cari_kod",
     None, None),
]

for sheet_name, df_kural, grp_kol, hata_kol, hata_esik in sekme_tanimi:
    df_to_sheet_grouped(wb, sheet_name, df_kural.reset_index(drop=True) if not df_kural.empty else df_kural,
                        gruplama_kol=grp_kol, miktar_hata_kol=hata_kol, hata_esik=hata_esik)
    print(f"  Sekme '{sheet_name}' yazıldı ({len(df_kural):,} satır)")

wb.save(CIKTI_EXCEL)
print(f"\nExcel kaydedildi: {CIKTI_EXCEL}")

# Doğrulama
wb2 = openpyxl.load_workbook(CIKTI_EXCEL, read_only=True)
print(f"Sekme sayısı: {len(wb2.sheetnames)} | Sekmeler: {wb2.sheetnames}")
wb2.close()
print("\nEXCEL RAPORU TAMAMLANDI ✓")
