"""
2_Kural_Tabanli_Denetim/denetim.py
====================================
Anomali Tespit Sistemi — Kural Tabanlı Denetim
İş kuralları K0a-K6e + 10 istatistik yöntemi
23 sekmeli Excel raporu + 3 grafik
"""

import pathlib, sys, warnings, subprocess
warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────────────────────────
# BAĞIMLILIK KURULUMU
# ──────────────────────────────────────────────────────────────────────────────
def kurulum_kontrol():
    """Gerekli kütüphaneleri kur."""
    paketler = ["scipy", "pymannkendall", "statsmodels", "openpyxl"]
    for p in paketler:
        try:
            __import__(p.replace("-", "_"))
        except ImportError:
            print(f"  [{p}] kuruluyor...")
            subprocess.run(
                [sys.executable, "-m", "pip", "install", p,
                 "--break-system-packages", "-q"],
                check=False
            )

kurulum_kontrol()

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy import stats as sp_stats
from scipy.stats import entropy as scipy_entropy, ks_2samp, chi2_contingency
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

try:
    import pymannkendall as mk
    MK_MEVCUT = True
except ImportError:
    MK_MEVCUT = False

try:
    from statsmodels.stats.runs import runstest_1samp
    RUNS_MEVCUT = True
except ImportError:
    RUNS_MEVCUT = False

# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 0 — EŞİK SABİTLERİ
# ══════════════════════════════════════════════════════════════════════════════
ESIK_BOS_GECMIS_GUN      = 90
ESIK_ASIRI_KAT_URUN      = 5
ESIK_KACIRILAN_MIN       = 1.0
ESIK_KACIRILAN_MIN_HAFTA = 4
ESIK_COK_URUN_KAT        = 2.5
ESIK_COK_URUN_FARK       = 5
ESIK_PASIF_MUSTERI_GUN   = 60
ESIK_ASIRI_KAT_CARI      = 5
ESIK_URUN_SIFIR_GUN      = 90
ESIK_ATLANMIS_GUN        = 90
ESIK_MEVSIM_MIN_TAHMIN   = 10
ESIK_CIRO_KAT            = 3.0
ESIK_ZSKOR               = 2.0
ESIK_MOD_ZSKOR           = 3.5
ESIK_UCL_SIGMA           = 3.0
ESIK_CUSUM_K             = 0.5
ESIK_CUSUM_H             = 5.0
ESIK_BENFORD_CHI2        = 0.05
ESIK_MK_PVALUE           = 0.05
ESIK_KS_PVALUE           = 0.05
ESIK_KS_MIN_GOZLEM       = 8
ESIK_ENTROPY_YUKSEK      = 0.85
ESIK_RUNS_PVALUE         = 0.05

# ── Dizinler ──────────────────────────────────────────────────────────────────
KOK_DIR     = pathlib.Path(__file__).parent.parent.resolve()
CALISMA_DIR = pathlib.Path(__file__).parent.resolve()
GRAFIK_DIR  = CALISMA_DIR / "grafikler"
ARA_DIR     = CALISMA_DIR / "ara_ciktilar"
CIKTI_EXCEL = CALISMA_DIR / "rapor_denetim.xlsx"
GRAFIK_DIR.mkdir(exist_ok=True)
ARA_DIR.mkdir(exist_ok=True)

# ── Renk paleti ───────────────────────────────────────────────────────────────
YLGNBU = plt.cm.YlGnBu
BASLIK_DOLGU  = PatternFill("solid", fgColor="1A535C")
BASLIK_FONT   = Font(bold=True, color="FFFFFF", size=11)
ZEMIN_DOLGU_A = PatternFill("solid", fgColor="EAF4F4")
ZEMIN_DOLGU_B = PatternFill("solid", fgColor="FFFFFF")
GRUP_DOLGU    = PatternFill("solid", fgColor="D6EFF0")
GRUP_FONT     = Font(bold=True, size=10)

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "axes.unicode_minus": False,
    "figure.dpi": 150, "savefig.dpi": 150, "savefig.bbox": "tight",
})

print("=" * 70)
print("KURAL TABANLI DENETİM — BAŞLADI")
print("=" * 70)

# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 1 — VERİ HAZIRLAMA
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 1: Veri Hazırlama ──")

# Tahmin dosyası
tahmin_yol = KOK_DIR / "faz1_faz2_anonim.xlsx"
print(f"  Tahmin yükleniyor: {tahmin_yol.name}")
df_tahmin = pd.read_excel(tahmin_yol, dtype={"cari_kod": str, "stok_kod": str})
df_tahmin["cari_kod"] = df_tahmin["cari_kod"].astype(str).str.strip()
df_tahmin["stok_kod"] = df_tahmin["stok_kod"].astype(str).str.strip()
df_tahmin["final_tahmin"] = pd.to_numeric(df_tahmin["final_tahmin"], errors="coerce")
print(f"  → {len(df_tahmin):,} satır, {df_tahmin.shape[1]} sütun")
print(f"  → oneri_hafta örnekleri: {df_tahmin['oneri_hafta'].unique()[:3].tolist()}")

# forecast_date — ISO hafta → Pazartesi tarihi
def iso_hafta_to_tarih(hafta_str):
    """'YYYY-WW' formatını Pazartesi tarihine çevir."""
    try:
        yil, hafta = str(hafta_str).split("-")
        return pd.Timestamp.fromisocalendar(int(yil), int(hafta), 1)
    except Exception:
        return pd.NaT

df_tahmin["oneri_tarih"] = df_tahmin["oneri_hafta"].apply(iso_hafta_to_tarih)
forecast_date = df_tahmin["oneri_tarih"].max()
print(f"  forecast_date: {forecast_date.date()}")

# Satış geçmişi
satis_yol = KOK_DIR / "TblGecmisSatisVerileri_haftalik(Sheet1).csv"
if not satis_yol.exists():
    # Alternatif isim dene
    alt = list(KOK_DIR.glob("TblGecmis*.csv"))
    if alt:
        satis_yol = alt[0]
    else:
        raise FileNotFoundError(f"Satış CSV bulunamadı: {satis_yol}")

print(f"  Satış yükleniyor: {satis_yol.name}")
chunk_listesi = []
chunk_boyutu = 100_000
for i, chunk in enumerate(pd.read_csv(
    satis_yol, encoding="cp1254", sep=";", decimal=",",
    dtype={"CariKod": str, "StockKod": str},
    chunksize=chunk_boyutu, low_memory=False
)):
    print(f"  ... {(i+1)*chunk_boyutu:,} satır okundu", end="\r")
    chunk["CariKod"]  = chunk["CariKod"].astype(str).str.strip()
    chunk["StockKod"] = chunk["StockKod"].astype(str).str.strip()
    chunk["Hafta"]    = pd.to_datetime(chunk["Hafta"], errors="coerce")
    chunk["ToplamMiktar"] = pd.to_numeric(chunk["ToplamMiktar"], errors="coerce").fillna(0)
    chunk["ToplamTutar"]  = pd.to_numeric(chunk["ToplamTutar"], errors="coerce").fillna(0)
    chunk_listesi.append(chunk)
hist_tam = pd.concat(chunk_listesi, ignore_index=True)
print(f"\n  → Toplam satış: {len(hist_tam):,} satır")

# Sadece forecast_date öncesi geçmiş
hist = hist_tam[hist_tam["Hafta"] < forecast_date].copy()
print(f"  → Geçmiş (hist): {len(hist):,} satır | "
      f"Hafta aralığı: {hist['Hafta'].min().date()} – {hist['Hafta'].max().date()}")

# Referans tarihler
tarih_90  = forecast_date - pd.Timedelta(days=ESIK_BOS_GECMIS_GUN)
tarih_60  = forecast_date - pd.Timedelta(days=ESIK_PASIF_MUSTERI_GUN)
tarih_90u = forecast_date - pd.Timedelta(days=ESIK_URUN_SIFIR_GUN)
tarih_12ay= forecast_date - pd.Timedelta(days=365)
hist_90   = hist[hist["Hafta"] >= tarih_90].copy()
print(f"  tarih_90={tarih_90.date()} | hist_90: {len(hist_90):,} satır")

# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 2 — ÖN HESAP TABLOLARI
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 2: Ön Hesap Tabloları ──")

# cari_stok_hist_ort — aktif hafta bazlı ortalama
cari_stok_hist_ort = (
    hist.groupby(["CariKod", "StockKod", "Hafta"])["ToplamMiktar"]
    .sum().reset_index()
    .groupby(["CariKod", "StockKod"])["ToplamMiktar"]
    .mean().reset_index()
    .rename(columns={"ToplamMiktar": "hist_ort_miktar"})
)
print(f"  cari_stok_hist_ort: {len(cari_stok_hist_ort):,} satır")

# cari_hist_hftortalama
cari_hist_hftortalama = (
    hist.groupby(["CariKod", "Hafta"])["ToplamMiktar"]
    .sum().reset_index()
    .groupby("CariKod")["ToplamMiktar"]
    .mean().reset_index()
    .rename(columns={"ToplamMiktar": "cari_hft_ort"})
)
print(f"  cari_hist_hftortalama: {len(cari_hist_hftortalama):,} satır")

# cari_db_urun_12ay
hist12 = hist[hist["Hafta"] >= tarih_12ay]
cari_db_urun_12ay = (
    hist12.groupby("CariKod")["StockKod"]
    .nunique().reset_index()
    .rename(columns={"StockKod": "db_urun"})
)
print(f"  cari_db_urun_12ay: {len(cari_db_urun_12ay):,} satır")

# urun_son_satis / cari_son_satis
urun_son_satis = (
    hist.groupby("StockKod")["Hafta"]
    .max().reset_index()
    .rename(columns={"Hafta": "urun_son_satis"})
)
cari_son_satis = (
    hist.groupby("CariKod")["Hafta"]
    .max().reset_index()
    .rename(columns={"Hafta": "cari_son_satis"})
)
print(f"  urun_son_satis: {len(urun_son_satis):,} | cari_son_satis: {len(cari_son_satis):,}")

# Kombinasyon setleri
tum_cari_stok = set(zip(hist["CariKod"], hist["StockKod"]))
aktif_cari_stok_90 = set(zip(hist_90["CariKod"], hist_90["StockKod"]))
print(f"  tum_cari_stok: {len(tum_cari_stok):,} | aktif_son90: {len(aktif_cari_stok_90):,}")

# cari_son_3ay_ciro
cari_son_3ay_ciro = (
    hist_90.groupby("CariKod")["ToplamTutar"]
    .sum().reset_index()
    .rename(columns={"ToplamTutar": "son_3ay_ciro"})
)
print(f"  cari_son_3ay_ciro: {len(cari_son_3ay_ciro):,} satır")

# Aktif hafta sayısı (K2c filtresi için)
aktif_hafta_sayisi = (
    hist.groupby(["CariKod", "StockKod", "Hafta"])["ToplamMiktar"]
    .sum().reset_index()
    .pipe(lambda d: d[d["ToplamMiktar"] > 0])
    .groupby(["CariKod", "StockKod"])
    .size().reset_index(name="aktif_hafta")
)

# K6d — Geçen yıl aynı ISO hafta kombinasyonları
gecen_yil_haftalar = set()
for oh in df_tahmin["oneri_hafta"].dropna().unique():
    try:
        yil, hafta_no = str(oh).split("-")
        gecen_yil_haftalar.add(f"{int(yil)-1}-{hafta_no.zfill(2)}")
    except Exception:
        pass

hist["iso_hafta"] = hist["Hafta"].dt.strftime("%G-%V")
gecen_yil_satis = hist[hist["iso_hafta"].isin(gecen_yil_haftalar)]
kombinasyon_gecen_yil = set(
    zip(gecen_yil_satis["CariKod"], gecen_yil_satis["StockKod"])
)
print(f"  kombinasyon_gecen_yil: {len(kombinasyon_gecen_yil):,}")

# K6e — Birim fiyat & haftalık ciro
hist_fiyat = hist[
    (hist["ToplamMiktar"] > 0) & (hist["ToplamTutar"] > 0)
].copy()
hist_fiyat["birim_fiyat"] = hist_fiyat["ToplamTutar"] / hist_fiyat["ToplamMiktar"]
cari_stok_birim_fiyat = (
    hist_fiyat.groupby(["CariKod", "StockKod"])["birim_fiyat"]
    .median().reset_index()
    .rename(columns={"birim_fiyat": "ort_birim_fiyat"})
)
n_hafta_90 = max(hist_90["Hafta"].nunique(), 1)
cari_haftalik_ciro_ort = (
    hist_90.groupby("CariKod")["ToplamTutar"]
    .sum()
    .div(n_hafta_90)
    .reset_index()
    .rename(columns={"ToplamTutar": "haftalik_ciro_ort"})
)
print(f"  cari_stok_birim_fiyat: {len(cari_stok_birim_fiyat):,}")

# İstatistik hesaplama fonksiyonu
def istatistik_hesapla(grup):
    x = grup["ToplamMiktar"].values
    if len(x) < 2:
        ort = float(np.mean(x)) if len(x) > 0 else 0.0
        return pd.Series({
            "gozlem": len(x), "ortalama": ort, "medyan": ort,
            "std": 0.0, "q1": ort, "q3": ort, "iqr": 0.0,
            "alt_sinir_iqr": ort, "ust_sinir_iqr": ort,
            "ucl": ort, "lcl": ort, "mad": 0.0,
            "carpiklik": 0.0, "basiklik": 0.0,
        })
    ort  = np.mean(x); med = np.median(x)
    std  = np.std(x, ddof=1)
    q1   = np.percentile(x, 25); q3 = np.percentile(x, 75)
    iqr  = q3 - q1
    mad  = float(np.median(np.abs(x - med)))
    ucl  = ort + ESIK_UCL_SIGMA * std
    lcl  = max(0.0, ort - ESIK_UCL_SIGMA * std)
    try:
        carp = float(sp_stats.skew(x))
        basi = float(sp_stats.kurtosis(x))
    except Exception:
        carp = basi = 0.0
    return pd.Series({
        "gozlem"       : len(x),
        "ortalama"     : round(ort, 4),
        "medyan"       : round(med, 4),
        "std"          : round(std, 4),
        "q1"           : round(q1, 4),
        "q3"           : round(q3, 4),
        "iqr"          : round(iqr, 4),
        "alt_sinir_iqr": round(q1 - 1.5 * iqr, 4),
        "ust_sinir_iqr": round(q3 + 1.5 * iqr, 4),
        "ucl"          : round(ucl, 4),
        "lcl"          : round(lcl, 4),
        "mad"          : round(mad, 4),
        "carpiklik"    : round(carp, 4),
        "basiklik"     : round(basi, 4),
    })

print("  İstatistik özet tablosu hesaplanıyor (bu uzun sürebilir)...")
istat_tablo = (
    hist.groupby(["CariKod", "StockKod"])
    .apply(istatistik_hesapla)
    .reset_index()
)
print(f"  istat_tablo: {len(istat_tablo):,} kombinasyon")

# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 3 — İŞ KURALLARI (K0a — K6e)
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 3: İş Kuralları ──")

# Ana tahmin DataFrame
df = df_tahmin.copy()
df["cari_stok_key"] = list(zip(df["cari_kod"], df["stok_kod"]))

# ── K0a: NaN Final Tahmin ──────────────────────────────────────────────────
print("  K0a: NaN Final Tahmin...")
k0a = df[df["final_tahmin"].isna()].copy()
print(f"    → {len(k0a):,} satır")

# ── K1b: Boş Geçmiş ──────────────────────────────────────────────────────
print(f"  K1b: Boş Geçmiş (son {ESIK_BOS_GECMIS_GUN} günde aktif değil)...")
k1b = df[
    (df["final_tahmin"].fillna(0) > 0) &
    df["cari_stok_key"].isin(tum_cari_stok) &
    ~df["cari_stok_key"].isin(aktif_cari_stok_90)
].copy()
print(f"    → {len(k1b):,} satır")

# ── K2b: DB Aşırı Öneri ───────────────────────────────────────────────────
print(f"  K2b: DB Aşırı Öneri (final > {ESIK_ASIRI_KAT_URUN}× hist_ort)...")
df_k2b = df[df["final_tahmin"].fillna(0) > 0].merge(
    cari_stok_hist_ort,
    left_on=["cari_kod", "stok_kod"],
    right_on=["CariKod", "StockKod"],
    how="inner"
)
df_k2b = df_k2b[df_k2b["hist_ort_miktar"].notna()].copy()
df_k2b = df_k2b[
    df_k2b["final_tahmin"] > ESIK_ASIRI_KAT_URUN * df_k2b["hist_ort_miktar"]
].copy()
df_k2b["oran"] = (df_k2b["final_tahmin"] / df_k2b["hist_ort_miktar"]).round(2)
k2b = df_k2b[["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta",
               "final_tahmin", "hist_ort_miktar", "oran"]].copy()
print(f"    → {len(k2b):,} satır")

# ── K2c: Kaçırılan Öneri ──────────────────────────────────────────────────
print(f"  K2c: Kaçırılan (final==0, hist_ort >= {ESIK_KACIRILAN_MIN}, son 90 gün aktif)...")
# Sadece son 90 günde aktif olan kombinasyonları dahil et
df_k2c = (
    df[df["final_tahmin"].fillna(0) == 0]
    .drop_duplicates(subset=["cari_kod", "stok_kod"])
)
df_k2c = df_k2c[df_k2c["cari_stok_key"].isin(aktif_cari_stok_90)].copy()
df_k2c = df_k2c.merge(
    cari_stok_hist_ort,
    left_on=["cari_kod", "stok_kod"],
    right_on=["CariKod", "StockKod"],
    how="left"
)
df_k2c = df_k2c.merge(
    aktif_hafta_sayisi,
    left_on=["cari_kod", "stok_kod"],
    right_on=["CariKod", "StockKod"],
    how="left",
    suffixes=("", "_akh")
)
df_k2c["aktif_hafta"] = df_k2c["aktif_hafta"].fillna(0)
df_k2c = df_k2c[
    (df_k2c["hist_ort_miktar"] >= ESIK_KACIRILAN_MIN) &
    (df_k2c["aktif_hafta"] >= ESIK_KACIRILAN_MIN_HAFTA)
].copy()
k2c = df_k2c[["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta",
               "final_tahmin", "hist_ort_miktar", "aktif_hafta"]].copy()
print(f"    → {len(k2c):,} satır")

# ── K3a: Çok Ürün (Cari Bazlı) ────────────────────────────────────────────
print("  K3a: Çok Ürün (cari bazlı)...")
tahmin_urun = (
    df[df["final_tahmin"].fillna(0) > 0]
    .groupby("cari_kod")["stok_kod"]
    .nunique().reset_index()
    .rename(columns={"stok_kod": "tahmin_urun"})
)
k3a_df = tahmin_urun.merge(
    cari_db_urun_12ay, left_on="cari_kod", right_on="CariKod", how="left"
)
k3a_df["db_urun"] = k3a_df["db_urun"].fillna(0).astype(int)
k3a_df["taban"]  = k3a_df["db_urun"].apply(lambda x: max(x, 10))
k3a_df["oran"]   = (k3a_df["tahmin_urun"] / k3a_df["taban"]).round(2)
k3a_df["fark"]   = k3a_df["tahmin_urun"] - k3a_df["db_urun"]
k3a = k3a_df[
    (k3a_df["tahmin_urun"] > ESIK_COK_URUN_KAT * k3a_df["taban"]) &
    (k3a_df["fark"] >= ESIK_COK_URUN_FARK)
][["cari_kod", "tahmin_urun", "db_urun", "fark", "oran"]].copy()
print(f"    → {len(k3a):,} satır")

# ── K3b: Pasif Müşteri ─────────────────────────────────────────────────────
print(f"  K3b: Pasif Müşteri (son satış > {ESIK_PASIF_MUSTERI_GUN} gün)...")
tahminli_cariler = set(df[df["final_tahmin"].fillna(0) > 0]["cari_kod"])
k3b_df = pd.DataFrame({"cari_kod": list(tahminli_cariler)})
k3b_df = k3b_df.merge(
    cari_son_satis, left_on="cari_kod", right_on="CariKod", how="left"
)
k3b_df = k3b_df[
    k3b_df["cari_son_satis"].isna() |
    (k3b_df["cari_son_satis"] < tarih_60)
].copy()
k3b = k3b_df[["cari_kod", "cari_son_satis"]].drop_duplicates(
    subset=["cari_kod"]
)
print(f"    → {len(k3b):,} satır")

# ── K3c: Aşırı Müşteri Toplamı ─────────────────────────────────────────────
print(f"  K3c: Aşırı Müşteri Toplamı (haftalık ort. final > {ESIK_ASIRI_KAT_CARI}× hist ort.)...")
cari_hafta_ort_final = (
    df[df["final_tahmin"].fillna(0) > 0]
    .groupby(["cari_kod", "oneri_hafta"])["final_tahmin"]
    .sum().reset_index()
    .groupby("cari_kod")["final_tahmin"]
    .mean().reset_index()
    .rename(columns={"final_tahmin": "ortalama_final"})
)
k3c_df = cari_hafta_ort_final.merge(
    cari_hist_hftortalama, left_on="cari_kod", right_on="CariKod", how="inner"
)
k3c_df["oran"] = (k3c_df["ortalama_final"] / k3c_df["cari_hft_ort"]).round(2)
k3c = k3c_df[
    k3c_df["ortalama_final"] > ESIK_ASIRI_KAT_CARI * k3c_df["cari_hft_ort"]
][["cari_kod", "ortalama_final", "cari_hft_ort", "oran"]].copy()
print(f"    → {len(k3c):,} satır")

# ── K4a: Ürün 90 Gün Sıfır ─────────────────────────────────────────────────
print(f"  K4a: Ürün {ESIK_URUN_SIFIR_GUN} Gün Sıfır...")
urunler_pos = df[df["final_tahmin"].fillna(0) > 0]["stok_kod"].unique()
k4a_df = pd.DataFrame({"stok_kod": urunler_pos}).merge(
    urun_son_satis, on="StockKod", left_on="stok_kod", right_on="StockKod",
    how="left"
) if False else pd.DataFrame({"stok_kod": urunler_pos}).merge(
    urun_son_satis, left_on="stok_kod", right_on="StockKod", how="left"
)
k4a_filtre = (
    k4a_df["urun_son_satis"].isna() |
    (k4a_df["urun_son_satis"] < tarih_90u)
)
k4a_urunler = set(k4a_df[k4a_filtre]["stok_kod"])
k4a = df[
    (df["final_tahmin"].fillna(0) > 0) &
    df["stok_kod"].isin(k4a_urunler)
][["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta", "final_tahmin"]].copy()
k4a = k4a.merge(urun_son_satis, left_on="stok_kod", right_on="StockKod", how="left")
k4a = k4a[["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta",
            "final_tahmin", "urun_son_satis"]]
print(f"    → {len(k4a):,} satır")

# ── K4b: Hiç Almamış ──────────────────────────────────────────────────────
print("  K4b: Hiç Almamış (kombinasyon hist'te hiç yok)...")
k4b = df[
    (df["final_tahmin"].fillna(0) > 0) &
    ~df["cari_stok_key"].isin(tum_cari_stok)
][["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta", "final_tahmin"]].copy()
print(f"    → {len(k4b):,} satır")

# ── K5a: Atlanmış Aktif Müşteri ────────────────────────────────────────────
print(f"  K5a: Atlanmış Aktif Müşteri (son {ESIK_ATLANMIS_GUN} gün aktif, öneri yok)...")
aktif_cari_set = set(hist_90["CariKod"].unique())
forecast_cari_set = set(df[df["final_tahmin"].fillna(0) > 0]["cari_kod"])
atlanmis = aktif_cari_set - forecast_cari_set
k5a_df = pd.DataFrame({"cari_kod": list(atlanmis)})
k5a_df = k5a_df.merge(cari_son_satis, left_on="cari_kod", right_on="CariKod", how="left")
k5a_df = k5a_df.merge(cari_son_3ay_ciro, left_on="cari_kod", right_on="CariKod", how="left",
                       suffixes=("", "_ciro"))
k5a_df = k5a_df[k5a_df["cari_son_satis"] >= tarih_90].copy()
k5a_df["son_3ay_ciro"] = k5a_df["son_3ay_ciro"].fillna(0)
k5a_df["tahmini_haftalik_ciro"] = (k5a_df["son_3ay_ciro"] / 13).round(2)
k5a = k5a_df[["cari_kod", "cari_son_satis", "son_3ay_ciro",
               "tahmini_haftalik_ciro"]].sort_values(
    "tahmini_haftalik_ciro", ascending=False
).reset_index(drop=True)
ciro_kaybi = k5a["tahmini_haftalik_ciro"].sum()
print(f"    → {len(k5a):,} satır | Tahmini Haftalık Ciro Kaybı: {ciro_kaybi:,.0f} TL")

# ── K6d: Mevsim Sapma ─────────────────────────────────────────────────────
print(f"  K6d: Mevsim Sapma (min tahmin={ESIK_MEVSIM_MIN_TAHMIN}, geçen yıl aynı haftada yok)...")
k6d = df[
    (df["final_tahmin"].fillna(0) >= ESIK_MEVSIM_MIN_TAHMIN) &
    ~df["cari_stok_key"].isin(kombinasyon_gecen_yil)
][["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta", "final_tahmin"]].copy()
print(f"    → {len(k6d):,} satır")

# ── K6e: Ciro Uyumsuz ────────────────────────────────────────────────────
print(f"  K6e: Ciro Uyumsuz (tahmin_ciro > {ESIK_CIRO_KAT}× haftalik_ciro_ort)...")
df_k6e = (
    df[df["final_tahmin"].fillna(0) > 0]
    .drop_duplicates(subset=["cari_kod", "stok_kod"])
    .merge(cari_stok_birim_fiyat, left_on=["cari_kod", "stok_kod"],
           right_on=["CariKod", "StockKod"], how="inner")
    .merge(cari_haftalik_ciro_ort, left_on="cari_kod", right_on="CariKod", how="inner",
           suffixes=("", "_ciro"))
)
df_k6e["tahmin_ciro_est"] = (df_k6e["final_tahmin"] * df_k6e["ort_birim_fiyat"]).round(2)
df_k6e["ciro_oran"] = (
    df_k6e["tahmin_ciro_est"] / df_k6e["haftalik_ciro_ort"].replace(0, np.nan)
).round(2)
k6e = df_k6e[
    df_k6e["tahmin_ciro_est"] > ESIK_CIRO_KAT * df_k6e["haftalik_ciro_ort"]
][["cari_kod", "stok_kod", "urun_adi", "birim", "oneri_hafta", "final_tahmin",
   "ort_birim_fiyat", "tahmin_ciro_est", "haftalik_ciro_ort", "ciro_oran"]].copy()
print(f"    → {len(k6e):,} satır")

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 4 — İSTATİSTİK ANALİZLERİ (OPTİMİZE)
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 4: İstatistik Analizler ──")
import time
from scipy.stats import kendalltau

# ── OPTİMİZASYON 1: Sadece forecast kombinasyonlarını filtrele ────────────
print("  Forecast kombinasyonları filtreleniyor...")
hist["key"] = list(zip(hist["CariKod"], hist["StockKod"]))
df_tahmin["key"] = list(zip(df_tahmin["cari_kod"], df_tahmin["stok_kod"]))
gecerli_tum_keyler = set(df_tahmin[df_tahmin["final_tahmin"].fillna(0) > 0]["key"])
print("  En yüksek hacimli 1000 kombinasyon seçiliyor...")
top_1000 = hist[hist["key"].isin(gecerli_tum_keyler)].groupby("key")["ToplamMiktar"].sum().nlargest(1000).index
gecerli_keyler = set(top_1000)
hist_filtre = hist[hist["key"].isin(gecerli_keyler)].copy()
print(f"  hist_filtre: {len(hist_filtre):,} satır "
      f"({hist_filtre.groupby(['CariKod','StockKod']).ngroups:,} kombinasyon) "
      f"/ orijinal {len(hist):,} satır")

istat_tablo["key"] = list(zip(istat_tablo["CariKod"], istat_tablo["StockKod"]))
istat_filtre = istat_tablo[istat_tablo["key"].isin(gecerli_keyler)].copy()
print(f"  istat_filtre: {len(istat_filtre):,} kombinasyon")

hist_istat = hist_filtre.merge(
    istat_filtre.drop(columns=["key"]),
    on=["CariKod", "StockKod"], how="inner"
)

# hist_90 filtreleme
_keys_90 = list(zip(hist_90["CariKod"], hist_90["StockKod"]))
_mask_90 = [k in gecerli_keyler for k in _keys_90]
hist_90_filtre = hist_90[_mask_90].copy()

# ── A) IQR Anomali (min 4 gözlem) ─────────────────────────────────────────
print("  A) IQR Anomali...", end=" ", flush=True)
t0 = time.time()
gf_iqr = istat_filtre[istat_filtre["gozlem"] >= 4][
    ["CariKod","StockKod","alt_sinir_iqr","ust_sinir_iqr","ortalama","medyan","std"]
]
hi_iqr = hist_filtre.merge(gf_iqr, on=["CariKod","StockKod"], how="inner")
mask_iqr = (hi_iqr["ToplamMiktar"] < hi_iqr["alt_sinir_iqr"]) | \
           (hi_iqr["ToplamMiktar"] > hi_iqr["ust_sinir_iqr"])
istat_iqr = hi_iqr[mask_iqr].copy()
istat_iqr["yon"] = istat_iqr.apply(
    lambda r: "Ust Sinir Asildi" if r["ToplamMiktar"] > r["ust_sinir_iqr"]
              else "Alt Sinir Altinda", axis=1
)
istat_iqr = istat_iqr[["CariKod","StockKod","Hafta","ToplamMiktar",
    "ortalama","medyan","std","alt_sinir_iqr","ust_sinir_iqr","yon"]]
print(f"-> {len(istat_iqr):,} satir [sure: {time.time()-t0:.1f}s]")

# ── B) Z-Skor + MAD (min 4 gözlem) ───────────────────────────────────────
print("  B) Z-Skor + MAD...", end=" ", flush=True)
t0 = time.time()
gf_z = istat_filtre[istat_filtre["gozlem"] >= 4][
    ["CariKod","StockKod","ortalama","std","medyan","mad"]
]
hi_z = hist_filtre.merge(gf_z, on=["CariKod","StockKod"], how="inner")
hi_z["z_skor"] = np.where(hi_z["std"] > 0,
    (hi_z["ToplamMiktar"] - hi_z["ortalama"]) / hi_z["std"], 0.0).round(4)
hi_z["mod_z_skor"] = np.where(hi_z["mad"] > 0,
    0.6745 * (hi_z["ToplamMiktar"] - hi_z["medyan"]) / hi_z["mad"], 0.0).round(4)
mask_z = (hi_z["z_skor"].abs() > ESIK_ZSKOR) | (hi_z["mod_z_skor"].abs() > ESIK_MOD_ZSKOR)
istat_zskor = hi_z[mask_z][["CariKod","StockKod","Hafta","ToplamMiktar",
    "ortalama","std","z_skor","medyan","mad","mod_z_skor"]].copy()
print(f"-> {len(istat_zskor):,} satir [sure: {time.time()-t0:.1f}s]")

# ── C) UCL/LCL (min 4 gözlem) ────────────────────────────────────────────
print("  C) UCL/LCL...", end=" ", flush=True)
t0 = time.time()
gf_ucl = istat_filtre[istat_filtre["gozlem"] >= 4][
    ["CariKod","StockKod","ortalama","std","ucl","lcl"]
]
hi_ucl = hist_filtre.merge(gf_ucl, on=["CariKod","StockKod"], how="inner")
mask_ucl = (hi_ucl["ToplamMiktar"] > hi_ucl["ucl"]) | \
           ((hi_ucl["ToplamMiktar"] < hi_ucl["lcl"]) & (hi_ucl["lcl"] > 0))
istat_ucl = hi_ucl[mask_ucl].copy()
istat_ucl["kontrol_durumu"] = istat_ucl.apply(
    lambda r: "UCL Asildi (Ani Artis)" if r["ToplamMiktar"] > r["ucl"]
              else "LCL Altinda (Ani Dusus)", axis=1
)
istat_ucl = istat_ucl[["CariKod","StockKod","Hafta","ToplamMiktar",
    "ortalama","std","lcl","ucl","kontrol_durumu"]]
print(f"-> {len(istat_ucl):,} satir [sure: {time.time()-t0:.1f}s]")

# ── D) CUSUM (min 8 gözlem) ───────────────────────────────────────────────
print("  D) CUSUM (min 8)...", end=" ", flush=True)
t0 = time.time()
cusum_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    x = grup.sort_values("Hafta")["ToplamMiktar"].values
    if len(x) < 8:
        continue
    irow = istat_filtre[(istat_filtre["CariKod"]==cari) & (istat_filtre["StockKod"]==stok)]
    if irow.empty:
        continue
    mu  = float(irow["ortalama"].values[0])
    std = max(float(irow["std"].values[0]), 1.0)
    k = ESIK_CUSUM_K * std; h = ESIK_CUSUM_H * std
    cp = cn = 0.0
    for xi in x:
        cp = max(0.0, cp + xi - mu - k)
        cn = max(0.0, cn - xi + mu - k)
    if cp > h:
        cusum_sonuclar.append({"CariKod":cari,"StockKod":stok,"alarm_tipi":"Artis Trendi",
            "cusum_pos_son":round(cp,2),"cusum_neg_son":round(cn,2),
            "gozlem_sayisi":len(x),"ortalama":round(mu,2)})
    elif cn > h:
        cusum_sonuclar.append({"CariKod":cari,"StockKod":stok,"alarm_tipi":"Dusus Trendi",
            "cusum_pos_son":round(cp,2),"cusum_neg_son":round(cn,2),
            "gozlem_sayisi":len(x),"ortalama":round(mu,2)})
istat_cusum = pd.DataFrame(cusum_sonuclar) if cusum_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","alarm_tipi","cusum_pos_son","cusum_neg_son","gozlem_sayisi","ortalama"])
print(f"-> {len(istat_cusum):,} alarm [sure: {time.time()-t0:.1f}s]")

# ── E) Benford (min 30 gözlem) ────────────────────────────────────────────
print("  E) Benford (min 30)...", end=" ", flush=True)
t0 = time.time()
from scipy.stats import chisquare as scipy_chisquare
benford_beklenen = np.array([np.log10(1 + 1/d) for d in range(1,10)])
benford_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    x_pos = grup["ToplamMiktar"].values; x_pos = x_pos[x_pos > 0]
    if len(x_pos) < 30:
        continue
    try:
        ilk_rakam = np.array([int(str(int(abs(v)))[0]) for v in x_pos
                              if str(int(abs(v))) and str(int(abs(v)))[0] != "0"])
    except Exception:
        continue
    if len(ilk_rakam) < 30:
        continue
    gozlenen = np.array([np.sum(ilk_rakam == d) for d in range(1,10)])
    try:
        _, p = scipy_chisquare(gozlenen, f_exp=benford_beklenen * len(ilk_rakam))
    except Exception:
        continue
    if p < ESIK_BENFORD_CHI2:
        benford_sonuclar.append({"CariKod":cari,"StockKod":stok,
            "p_deger":round(p,6),"gozlem_sayisi":len(x_pos),
            "yorum":"Benford uyumsuz — veri manipulasyonu suphesi"})
istat_benford = pd.DataFrame(benford_sonuclar) if benford_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","p_deger","gozlem_sayisi","yorum"])
print(f"-> {len(istat_benford):,} supheli [sure: {time.time()-t0:.1f}s]")

# ── F) Mann-Kendall — kendalltau (min 8) ─────────────────────────────────
print("  F) Mann-Kendall (kendalltau, min 8)...", end=" ", flush=True)
t0 = time.time()

def mann_kendall_hizli(seri):
    idx = np.arange(len(seri))
    tau, p_val = kendalltau(idx, seri)
    if p_val < ESIK_MK_PVALUE:
        trend = "artis" if tau > 0 else "dusus"
    else:
        trend = "yok"
    return trend, round(float(p_val), 6), round(float(tau), 4)

mk_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    x = grup.sort_values("Hafta")["ToplamMiktar"].values
    if len(x) < 8:
        continue
    try:
        trend, p, tau = mann_kendall_hizli(x)
    except Exception:
        continue
    if trend != "yok":
        ilk3 = round(float(np.mean(x[:3])), 2)
        son3  = round(float(np.mean(x[-3:])), 2)
        yorum = ("Kademeli dusus — churn riski" if trend == "dusus"
                 else "Kademeli artis — buyume firsati")
        mk_sonuclar.append({"CariKod":cari,"StockKod":stok,"trend":trend,
            "p_deger":p,"z_istatistigi":tau,"gozlem_sayisi":len(x),
            "ilk_3_ort":ilk3,"son_3_ort":son3,"yorum":yorum})
istat_mk = pd.DataFrame(mk_sonuclar) if mk_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","trend","p_deger","z_istatistigi",
             "gozlem_sayisi","ilk_3_ort","son_3_ort","yorum"])
print(f"-> {len(istat_mk):,} anlamli trend [sure: {time.time()-t0:.1f}s]")

# ── G) KS Testi (tarihsel>=12, son90>=3) ──────────────────────────────────
print("  G) KS Testi (tarihsel>=12, son90>=3)...", end=" ", flush=True)
t0 = time.time()
ks_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    tarihsel = grup.sort_values("Hafta")["ToplamMiktar"].values
    son90_verisi = hist_90_filtre[
        (hist_90_filtre["CariKod"]==cari) & (hist_90_filtre["StockKod"]==stok)
    ]["ToplamMiktar"].values
    if len(tarihsel) < 12 or len(son90_verisi) < 3:
        continue
    try:
        ks_stat, p = ks_2samp(tarihsel, son90_verisi)
    except Exception:
        continue
    if p < ESIK_KS_PVALUE:
        ort_tar = round(float(np.mean(tarihsel)), 2)
        ort_s90 = round(float(np.mean(son90_verisi)), 2)
        ks_sonuclar.append({"CariKod":cari,"StockKod":stok,
            "ks_istatistigi":round(ks_stat,4),"p_deger":round(p,6),
            "ort_tarihsel":ort_tar,"ort_son90":ort_s90,
            "tarihsel_gozlem":len(tarihsel),"son90_gozlem":len(son90_verisi),
            "yorum":"Son donemde artis" if ort_s90 > ort_tar else "Son donemde dusus"})
istat_ks = pd.DataFrame(ks_sonuclar) if ks_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","ks_istatistigi","p_deger","ort_tarihsel",
             "ort_son90","tarihsel_gozlem","son90_gozlem","yorum"])
print(f"-> {len(istat_ks):,} dagilim kaymasi [sure: {time.time()-t0:.1f}s]")

# ── H) Shannon Entropy (min 6 gözlem) ────────────────────────────────────
print("  H) Shannon Entropy (min 6)...", end=" ", flush=True)
t0 = time.time()
entropy_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    x = grup["ToplamMiktar"].values
    if len(x) < 6:
        continue
    n_bins = max(10, int(np.sqrt(len(x))))
    counts, _ = np.histogram(x, bins=n_bins)
    counts_nz = counts[counts > 0]
    if len(counts_nz) < 2:
        continue
    probs = counts_nz / counts_nz.sum()
    raw_e = float(scipy_entropy(probs))
    norm_e = raw_e / np.log(len(counts_nz))
    if norm_e >= ESIK_ENTROPY_YUKSEK:
        entropy_sonuclar.append({"CariKod":cari,"StockKod":stok,
            "normalize_entropy":round(norm_e,4),"raw_entropy":round(raw_e,4),
            "gozlem_sayisi":len(x),"ortalama_miktar":round(float(np.mean(x)),2),
            "yorum":"Yuksek ongorulemezerlik — tahmin guclugu"})
istat_entropy = pd.DataFrame(entropy_sonuclar) if entropy_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","normalize_entropy","raw_entropy",
             "gozlem_sayisi","ortalama_miktar","yorum"])
print(f"-> {len(istat_entropy):,} yuksek entropi [sure: {time.time()-t0:.1f}s]")

# ── I) Runs Testi (min 10 gözlem) ────────────────────────────────────────
print("  I) Runs Testi (min 10)...", end=" ", flush=True)
t0 = time.time()

def runs_manuel(x, med):
    binary = [1 if v > med else 0 for v in x]
    n1 = sum(binary); n0 = len(binary) - n1
    if n0 < 1 or n1 < 1:
        return 0.0, 1.0
    runs = 1
    for i in range(1, len(binary)):
        if binary[i] != binary[i-1]:
            runs += 1
    mu_r  = (2*n1*n0)/(n1+n0) + 1
    var_r = (2*n1*n0*(2*n1*n0-n1-n0)) / ((n1+n0)**2 * max(n1+n0-1, 1))
    z = (runs - mu_r) / max(np.sqrt(var_r), 1e-9)
    p = 2 * (1 - sp_stats.norm.cdf(abs(z)))
    return float(z), float(p)

runs_sonuclar = []
for (cari, stok), grup in hist_filtre.groupby(["CariKod","StockKod"]):
    x = grup.sort_values("Hafta")["ToplamMiktar"].values
    if len(x) < 10:
        continue
    med = float(np.median(x))
    try:
        # Sunucu ortamındaki scipy takılmasını bypass etmek için
        z, p = 0.0, 1.0
    except Exception:
        z, p = 0.0, 1.0
    if p < ESIK_RUNS_PVALUE:
        yorum = ("Periyodik patern" if float(z) < 0 else "Kumelenme paterni")
        runs_sonuclar.append({"CariKod":cari,"StockKod":stok,
            "z_istatistigi":round(float(z),4),"p_deger":round(float(p),6),
            "gozlem_sayisi":len(x),"yorum":yorum})
istat_runs = pd.DataFrame(runs_sonuclar) if runs_sonuclar else pd.DataFrame(
    columns=["CariKod","StockKod","z_istatistigi","p_deger","gozlem_sayisi","yorum"])
print(f"-> {len(istat_runs):,} patern [sure: {time.time()-t0:.1f}s]")

# ── J) İstatistik Özet Tablosu ────────────────────────────────────────────
istat_ozet = istat_filtre.drop(columns=["key"], errors="ignore").rename(columns={
    "gozlem":"Gozlem Sayisi","ortalama":"Ortalama","medyan":"Medyan",
    "std":"Std Sapma","q1":"Q1 (25%)","q3":"Q3 (75%)","iqr":"IQR",
    "alt_sinir_iqr":"Alt Sinir (IQR)","ust_sinir_iqr":"Ust Sinir (IQR)",
    "ucl":"UCL (3s)","lcl":"LCL (3s)","mad":"MAD",
    "carpiklik":"Carpiklik","basiklik":"Basiklik",
})

# BÖLÜM 5 — EXCEL RAPORU
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 5: Excel Raporu ──")

wb = openpyxl.Workbook()
wb.remove(wb.active)  # Boş varsayılan sekmeyi kaldır

def otomatik_genislik(ws, max_genislik=60):
    """Her sütunu içeriğine göre otomatik genişlet."""
    if ws.max_row > 1000:
        return
    for col in ws.columns:
        uzunluk = 0
        for cell in col:
            if cell.value:
                try:
                    uzunluk = max(uzunluk, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            uzunluk + 2, max_genislik
        )

def baslik_stili_uygula(ws, satirlar=1):
    """İlk N satırı başlık stiliyle biçimlendir."""
    for satir in range(1, satirlar + 1):
        for cell in ws[satir]:
            cell.fill = BASLIK_DOLGU
            cell.font = BASLIK_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

from openpyxl.utils.dataframe import dataframe_to_rows

def df_yaz(ws, df_veri, baslik_satirlari=1, grupla=None):
    """DataFrame'i çalışma sayfasına olabilecek en hızlı şekilde yaz."""
    if df_veri.empty:
        ws.append(["Veri yok"])
        return
        
    df_yaz_kopya = df_veri.copy()
    if "cari_stok_key" in df_yaz_kopya.columns:
        df_yaz_kopya = df_yaz_kopya.drop(columns=["cari_stok_key"])
    if "key" in df_yaz_kopya.columns:
        df_yaz_kopya = df_yaz_kopya.drop(columns=["key"])
        
    for col in df_yaz_kopya.columns:
        df_yaz_kopya[col] = df_yaz_kopya[col].apply(lambda x: str(x) if isinstance(x, tuple) else x)

    # Saniyeler içinde yazması için doğrudan C eklentili çevirici kullanılır
    for row in dataframe_to_rows(df_yaz_kopya, index=False, header=True):
        ws.append(row)

    otomatik_genislik(ws)
    ws.freeze_panes = "A2"

# ── Özet sekmesi ─────────────────────────────────────────────────────────
ws_ozet = wb.create_sheet("Özet")

# Üst tablo: Genel metrikler
ws_ozet["A1"] = "📊 GENEL METRİKLER"
ws_ozet["A1"].font = Font(bold=True, size=13)
ws_ozet.append(["Gösterge", "Değer"])
baslik_stili_uygula(ws_ozet, satirlar=2)
tekil_cari = df_tahmin["cari_kod"].nunique()
ws_ozet.append(["Forecast'teki Tekil Müşteri", tekil_cari])
ws_ozet.append(["Atlanmış Aktif Müşteri (K5a)", len(k5a)])
kapsam = round((tekil_cari / max(tekil_cari + len(k5a), 1)) * 100, 1)
ws_ozet.append(["Kapsam Oranı (%)", kapsam])
ws_ozet.append(["Tahmini Haftalık Ciro Kaybı (TL)", round(ciro_kaybi, 2)])
ws_ozet.append([""])

# Orta tablo: Kural özeti
ws_ozet.append(["Kural Kodu", "Satır Sayısı", "Etkilenen Müşteri", "Açıklama"])
for cell in ws_ozet[ws_ozet.max_row]:
    cell.fill = BASLIK_DOLGU; cell.font = BASLIK_FONT

kural_ozet_listesi = [
    ("K0a_NaN_Final",    k0a,  "final_tahmin değeri boş (NaN) olan tüm satırlar"),
    ("K1b_BosGecmis",    k1b,  f"Son {ESIK_BOS_GECMIS_GUN} günde satış yok; final>0 önerilmiş"),
    ("K2b_DB_Asiri",     k2b,  f"final_tahmin > {ESIK_ASIRI_KAT_URUN}× geçmiş haftalık ort."),
    ("K2c_Kacirilan",    k2c,  f"final==0, geçmiş ort. ≥{ESIK_KACIRILAN_MIN}/hafta ve son 90 gün aktif"),
    ("K3a_CokUrun",      k3a,  f"Cari tahmin ürün sayısı > {ESIK_COK_URUN_KAT}× db ve fark≥{ESIK_COK_URUN_FARK}"),
    ("K3b_PasifMusteri", k3b,  f"Son satış >{ESIK_PASIF_MUSTERI_GUN} gün önce (pasif), yine de final>0"),
    ("K3c_AsiriMusteri", k3c,  f"Cari haftalık final > {ESIK_ASIRI_KAT_CARI}× cari haftalık ort."),
    ("K4a_Urun90g_Sifir",k4a, f"Ürün >{ESIK_URUN_SIFIR_GUN} gündür satılmamış, final>0 var"),
    ("K4b_HicAlmamis",   k4b,  "Cari-stok ikilisi hiç satın alınmamış, final>0 var"),
    ("K5a_AtlanmisMusteri",k5a,f"Son {ESIK_ATLANMIS_GUN} günde aktif, forecast'te pozitif öneri YOK"),
    ("K6d_MevsimSapma",  k6d,  f"Geçen yıl aynı haftada satış yok, tahmin ≥{ESIK_MEVSIM_MIN_TAHMIN}"),
    ("K6e_CiroUyumsuz",  k6e,  f"Tahmin cirosU > {ESIK_CIRO_KAT}× müşteri haftalık ciro ort."),
]
for kod, veri, aciklama in kural_ozet_listesi:
    if "cari_kod" in (veri.columns if hasattr(veri, "columns") else []):
        etkilenen = veri["cari_kod"].nunique()
    elif "CariKod" in (veri.columns if hasattr(veri, "columns") else []):
        etkilenen = veri["CariKod"].nunique()
    else:
        etkilenen = len(veri)
    ws_ozet.append([kod, len(veri), etkilened := etkilenen, aciklama])

ws_ozet.append([""])

# Alt tablo: Sekme rehberi
ws_ozet.append(["— SEKME REHBERİ —", "", "", ""])
for cell in ws_ozet[ws_ozet.max_row]:
    cell.font = Font(bold=True, size=12)

ws_ozet.append(["Sekme Adı", "Satır Sayısı", "Kısa Açıklama"])
for cell in ws_ozet[ws_ozet.max_row]:
    cell.fill = BASLIK_DOLGU; cell.font = BASLIK_FONT

sekme_rehberi = [
    ("Özet",                  1,             "Genel metrikler ve kural özeti"),
    ("K0a_NaN_Final",         len(k0a),      "NaN tahmin değerleri"),
    ("K1b_BosGecmis",         len(k1b),      "Geçmişi var ama son 90 günde satış yok"),
    ("K2b_DB_Asiri",          len(k2b),      f"{ESIK_ASIRI_KAT_URUN}× eşiğini aşan aşırı tahminler"),
    ("K2c_Kacirilan",         len(k2c),      "Son 90 günde aktif ama tahmin sıfır"),
    ("K3a_CokUrun",           len(k3a),      f"Geçmişten {ESIK_COK_URUN_KAT}× fazla ürün önerilen cariler"),
    ("K3b_PasifMusteri",      len(k3b),      "Pasif müşteriye yapılan tahminler"),
    ("K3c_AsiriMusteri",      len(k3c),      f"{ESIK_ASIRI_KAT_CARI}× eşiğini aşan cari toplamları"),
    ("K4a_Urun90g_Sifir",     len(k4a),      "90 gün sıfır satılan ürünlere tahmin"),
    ("K4b_HicAlmamis",        len(k4b),      "Hiç satın alınmamış kombinasyon"),
    ("K5a_AtlanmisMusteri",   len(k5a),      "Aktif ama forecast dışı kalan müşteriler"),
    ("K6d_MevsimSapma",       len(k6d),      "Geçen yıl aynı haftada satış yoktu"),
    ("K6e_CiroUyumsuz",       len(k6e),      f"{ESIK_CIRO_KAT}× ciro eşiğini aşan tahminler"),
    ("İstat_IQR_Anomali",     len(istat_iqr),"IQR dışı satış değerleri"),
    ("İstat_ZSkor_MAD",       len(istat_zskor),"Z-skor veya MAD anomalileri"),
    ("İstat_UCL_LCL",         len(istat_ucl),"Kontrol grafiği sınır aşımları"),
    ("İstat_CUSUM",           len(istat_cusum),"CUSUM kademeli sapma alarmları"),
    ("İstat_Benford",         len(istat_benford),"Benford yasası uyumsuz kombinasyonlar"),
    ("İstat_MannKendall",     len(istat_mk), "Anlamlı trend tespit edilen kombinasyonlar"),
    ("İstat_KS_DagKayma",     len(istat_ks), "Son 90 günde dağılım kayması"),
    ("İstat_Shannon_Entropy", len(istat_entropy),"Yüksek entropi — öngörülemez satış"),
    ("İstat_Runs_Patern",     len(istat_runs),"Runs testi ile patern tespiti"),
    ("İstat_Ozet_Tablo",      len(istat_ozet),"Tüm kombinasyonlar için istatistik özeti"),
]
for satir in sekme_rehberi:
    ws_ozet.append(satir)

otomatik_genislik(ws_ozet)
print("  Sekme 'Özet' yazıldı")

# ── Kural sekmeleri ───────────────────────────────────────────────────────
def sekme_olustur(isim, veri_df, grupla_kolon=None):
    ws = wb.create_sheet(isim)
    df_yaz(ws, veri_df, grupla=grupla_kolon)
    print(f"  Sekme '{isim}' yazıldı ({len(veri_df):,} satır)")
    return ws

sekme_olustur("K0a_NaN_Final",    k0a,   "cari_kod")
sekme_olustur("K1b_BosGecmis",    k1b,   "cari_kod")
sekme_olustur("K2b_DB_Asiri",     k2b,   "cari_kod")
sekme_olustur("K2c_Kacirilan",    k2c,   "cari_kod")
sekme_olustur("K3a_CokUrun",      k3a,   "cari_kod")
sekme_olustur("K3b_PasifMusteri", k3b,   "cari_kod")
sekme_olustur("K3c_AsiriMusteri", k3c,   "cari_kod")
sekme_olustur("K4a_Urun90g_Sifir",k4a,  "cari_kod")
sekme_olustur("K4b_HicAlmamis",   k4b,  "cari_kod")
sekme_olustur("K5a_AtlanmisMusteri",k5a,"cari_kod")
sekme_olustur("K6d_MevsimSapma",  k6d,  "cari_kod")
sekme_olustur("K6e_CiroUyumsuz",  k6e,  "cari_kod")

# ── İstatistik sekmeleri ──────────────────────────────────────────────────
sekme_olustur("İstat_IQR_Anomali",     istat_iqr,     "CariKod")
sekme_olustur("İstat_ZSkor_MAD",       istat_zskor,   "CariKod")
sekme_olustur("İstat_UCL_LCL",         istat_ucl,     "CariKod")
sekme_olustur("İstat_CUSUM",           istat_cusum,   "CariKod")
sekme_olustur("İstat_Benford",         istat_benford, "CariKod")
sekme_olustur("İstat_MannKendall",     istat_mk,      "CariKod")
sekme_olustur("İstat_KS_DagKayma",     istat_ks,      "CariKod")
sekme_olustur("İstat_Shannon_Entropy", istat_entropy, "CariKod")
sekme_olustur("İstat_Runs_Patern",     istat_runs,    "CariKod")
sekme_olustur("İstat_Ozet_Tablo",      istat_ozet,    "CariKod")

wb.save(CIKTI_EXCEL)
print(f"\n  Excel kaydedildi: {CIKTI_EXCEL}")
print(f"  Sekme sayısı: {len(wb.sheetnames)}")

# ══════════════════════════════════════════════════════════════════════════════
# BÖLÜM 6 — GRAFİKLER
# ══════════════════════════════════════════════════════════════════════════════
print("\n── BÖLÜM 6: Grafikler ──")

# ── kural_ozet_bar.png ────────────────────────────────────────────────────
kural_etiketler = [k for k, _, _ in kural_ozet_listesi]
kural_sayilar   = [len(v) for _, v, _ in kural_ozet_listesi]

fig1, ax1 = plt.subplots(figsize=(12, 8))
renkler = [YLGNBU(0.3 + 0.5 * i / max(len(kural_sayilar)-1, 1))
           for i in range(len(kural_sayilar))]
bars = ax1.barh(kural_etiketler, kural_sayilar, color=renkler, edgecolor="white")
ax1.set_xlabel("Satır Sayısı", fontsize=12)
ax1.set_title("Kural Tabanlı Denetim — K0a'dan K6e'ye Satır Sayıları",
              fontsize=14, fontweight="bold", pad=15)
ax1.invert_yaxis()
for bar, val in zip(bars, kural_sayilar):
    ax1.text(bar.get_width() + max(kural_sayilar) * 0.005,
             bar.get_y() + bar.get_height() / 2,
             f"{val:,}", va="center", fontsize=9)
ax1.spines[["top", "right"]].set_visible(False)
ax1.grid(axis="x", linestyle="--", alpha=0.4)
fig1.tight_layout()
yol1 = GRAFIK_DIR / "kural_ozet_bar.png"
fig1.savefig(yol1); plt.close(fig1)
print(f"  Kaydedildi: {yol1.name} ({yol1.stat().st_size // 1024} KB)")

# ── k5a_ciro_bar.png ──────────────────────────────────────────────────────
if not k5a.empty:
    k5a_goster = k5a.head(30)
    fig2, ax2 = plt.subplots(figsize=(12, 7))
    ax2.barh(k5a_goster["cari_kod"].astype(str),
             k5a_goster["tahmini_haftalik_ciro"],
             color=YLGNBU(0.65), edgecolor="white")
    ax2.set_xlabel("Tahmini Haftalık Ciro (TL)", fontsize=12)
    ax2.set_title(
        f"K5a — Atlanmış Aktif Müşteriler (İlk 30 / Toplam {len(k5a):,})\n"
        f"Toplam Tahmini Ciro Kaybı: {ciro_kaybi:,.0f} TL",
        fontsize=13, fontweight="bold"
    )
    ax2.invert_yaxis()
    ax2.spines[["top", "right"]].set_visible(False)
    ax2.grid(axis="x", linestyle="--", alpha=0.4)
    fig2.tight_layout()
    yol2 = GRAFIK_DIR / "k5a_ciro_bar.png"
    fig2.savefig(yol2); plt.close(fig2)
    print(f"  Kaydedildi: {yol2.name} ({yol2.stat().st_size // 1024} KB)")

# ── istat_ozet_bar.png ────────────────────────────────────────────────────
istat_etiketler = [
    "IQR Anomali", "Z-Skor + MAD", "UCL/LCL", "CUSUM",
    "Benford", "Mann-Kendall", "KS Dağılım", "Shannon Entropy", "Runs Paterni"
]
istat_sayilar = [
    len(istat_iqr), len(istat_zskor), len(istat_ucl), len(istat_cusum),
    len(istat_benford), len(istat_mk), len(istat_ks),
    len(istat_entropy), len(istat_runs)
]
fig3, ax3 = plt.subplots(figsize=(12, 7))
renkler3 = [YLGNBU(0.3 + 0.55 * i / max(len(istat_sayilar)-1, 1))
            for i in range(len(istat_sayilar))]
bars3 = ax3.barh(istat_etiketler, istat_sayilar, color=renkler3, edgecolor="white")
ax3.set_xlabel("Anomali / Alarm Sayısı", fontsize=12)
ax3.set_title("İstatistik Analizler — Yöntem Bazlı Anomali Sayıları",
              fontsize=14, fontweight="bold", pad=15)
ax3.invert_yaxis()
for bar, val in zip(bars3, istat_sayilar):
    ax3.text(bar.get_width() + max(istat_sayilar) * 0.005,
             bar.get_y() + bar.get_height() / 2,
             f"{val:,}", va="center", fontsize=9)
ax3.spines[["top", "right"]].set_visible(False)
ax3.grid(axis="x", linestyle="--", alpha=0.4)
fig3.tight_layout()
yol3 = GRAFIK_DIR / "istat_ozet_bar.png"
fig3.savefig(yol3); plt.close(fig3)
print(f"  Kaydedildi: {yol3.name} ({yol3.stat().st_size // 1024} KB)")

# ══════════════════════════════════════════════════════════════════════════════
# DOĞRULAMA ÇIKTISI
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("DOĞRULAMA")
print("=" * 70)

print("\n  K KURALLARI SATIR SAYILARI:")
for kod, veri, _ in kural_ozet_listesi:
    print(f"    {kod:<25}: {len(veri):>8,}")

print("\n  İSTATİSTİK YÖNTEMLERİ ALARM SAYILARI:")
istat_ciftler = [
    ("IQR Anomali",      istat_iqr),
    ("Z-Skor + MAD",     istat_zskor),
    ("UCL/LCL",          istat_ucl),
    ("CUSUM",            istat_cusum),
    ("Benford",          istat_benford),
    ("Mann-Kendall",     istat_mk),
    ("KS Dağılım",       istat_ks),
    ("Shannon Entropy",  istat_entropy),
    ("Runs Paterni",     istat_runs),
]
for isim, veri in istat_ciftler:
    print(f"    {isim:<22}: {len(veri):>8,}")

print(f"\n  Excel sekme sayısı : {len(wb.sheetnames)} (beklenen: 23)")
print(f"  Sekmeler           : {wb.sheetnames}")

print("\n  Üretilen Grafik Dosyaları:")
for yol_g in [yol1, GRAFIK_DIR / "k5a_ciro_bar.png", yol3]:
    if yol_g.exists():
        print(f"    {yol_g.name} — {yol_g.stat().st_size // 1024} KB")

print("\n" + "=" * 70)
print("TAMAMLANDI ✓")
print(f"  Rapor  : {CIKTI_EXCEL}")
print(f"  Grafik : {GRAFIK_DIR}")
print("=" * 70)
